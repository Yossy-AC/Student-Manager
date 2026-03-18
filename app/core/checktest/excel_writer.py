"""Excel出力"""

import logging
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .constants import COLOR_FAIL, COLOR_FLAG, COLOR_HEADER, COLOR_MULTI

log = logging.getLogger(__name__)


def write_excel(results: list, config: dict, output_path: str) -> None:
    """処理結果を Excel ファイルに出力する"""
    wb = openpyxl.Workbook()
    questions = config["questions"]
    n_q = len(questions)

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    flag_fill   = PatternFill("solid", fgColor=COLOR_FLAG)
    multi_fill  = PatternFill("solid", fgColor=COLOR_MULTI)
    fail_fill   = PatternFill("solid", fgColor=COLOR_FAIL)

    headers = ["ページ", "氏名"] + [q["label"] for q in questions] + ["合計", "フラグ"]
    total_col = 3 + n_q
    flag_col  = total_col + 1

    def write_header(ws):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def set_col_widths(ws):
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 18
        for qi in range(n_q):
            ws.column_dimensions[openpyxl.utils.get_column_letter(3 + qi)].width = 10
        ws.column_dimensions[openpyxl.utils.get_column_letter(total_col)].width = 8
        ws.column_dimensions[openpyxl.utils.get_column_letter(flag_col)].width = 22

    # ===== 「結果」シート =====
    ws = wb.active
    ws.title = "結果"
    write_header(ws)

    for ri, r in enumerate(results, 2):
        ws.cell(row=ri, column=1, value=r["page"])
        ws.cell(row=ri, column=2, value=r["name"])

        is_fail = r["page_flag"] == "読み取り失敗"
        row_fill = fail_fill if is_fail else (flag_fill if r["page_flag"] != "正常" else None)

        for qi, score in enumerate(r["scores"]):
            ci = 3 + qi
            cell = ws.cell(row=ri, column=ci, value=score)
            cell.alignment = Alignment(horizontal="center")
            if r["flags"][qi] == "複数塗り":
                cell.fill = multi_fill
            elif row_fill:
                cell.fill = row_fill

        tc = ws.cell(row=ri, column=total_col, value=r["total_mark"])
        tc.alignment = Alignment(horizontal="center")
        if row_fill:
            tc.fill = row_fill

        fc = ws.cell(row=ri, column=flag_col, value=r["page_flag"])
        if row_fill:
            fc.fill = row_fill

    # 集計行（最終行の下）
    n_data = len(results)
    if n_data > 0:
        data_end_row = n_data + 1
        for label, summary_row in [("平均", n_data + 3), ("最高", n_data + 4), ("最低", n_data + 5)]:
            ws.cell(row=summary_row, column=2, value=label).font = Font(bold=True)
            for qi in range(n_q):
                ci = 3 + qi
                col_l = openpyxl.utils.get_column_letter(ci)
                ref = f"{col_l}2:{col_l}{data_end_row}"
                if label == "平均":
                    ws.cell(row=summary_row, column=ci, value=f"=AVERAGE({ref})")
                elif label == "最高":
                    ws.cell(row=summary_row, column=ci, value=f"=MAX({ref})")
                else:
                    ws.cell(row=summary_row, column=ci, value=f"=MIN({ref})")
            # 合計列
            tl = openpyxl.utils.get_column_letter(total_col)
            tref = f"{tl}2:{tl}{data_end_row}"
            if label == "平均":
                ws.cell(row=summary_row, column=total_col, value=f"=AVERAGE({tref})")
            elif label == "最高":
                ws.cell(row=summary_row, column=total_col, value=f"=MAX({tref})")
            else:
                ws.cell(row=summary_row, column=total_col, value=f"=MIN({tref})")

    set_col_widths(ws)

    # ===== 「要確認」シート =====
    ws2 = wb.create_sheet("要確認")
    write_header(ws2)
    row2 = 2
    for r in results:
        if r["page_flag"] not in ("正常",):
            for ci, val in enumerate(
                [r["page"], r["name"]] + r["scores"] + [r["total_mark"], r["page_flag"]], 1
            ):
                cell = ws2.cell(row=row2, column=ci, value=val)
                if 3 <= ci < 3 + len(r["flags"]) and r["flags"][ci - 3] == "複数塗り":
                    cell.fill = multi_fill
            row2 += 1
    set_col_widths(ws2)

    # 保存
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    log.info(f"Excel 保存: {output_path}")
